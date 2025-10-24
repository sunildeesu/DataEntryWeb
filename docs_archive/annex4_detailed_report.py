import openpyxl
from openpyxl.utils import get_column_letter
import json

def rgb_to_hex(rgb):
    """Convert RGB to hex color string"""
    if rgb and len(str(rgb)) >= 6:
        return f"#{str(rgb)[-6:]}"
    return None

print("="*100)
print("DETAILED ANNEX-IV (ANEX-IV) SHEET ANALYSIS")
print("="*100)

# Load workbook
wb = openpyxl.load_workbook('/Users/sunildeesu/myProjects/Hemanth/DataEntryWeb/QAQC-FS8-Acceptance.xlsx', data_only=False)
sheet = wb['ANEX-IV']

print(f"\nSheet Name: {sheet.title}")
print(f"Max Row: {sheet.max_row}")
print(f"Max Column: {sheet.max_column} ({get_column_letter(sheet.max_column)})")

# 1. HEADER ROW (Row 1) - COMPLETE ANALYSIS
print(f"\n{'='*100}")
print("1. HEADER ROW (Row 1) - COMPLETE DETAILS")
print(f"{'='*100}")

headers = []
for col_num in range(1, 17):  # Check first 16 columns
    cell = sheet.cell(row=1, column=col_num)
    col_letter = get_column_letter(col_num)

    header_info = {
        'column': col_letter,
        'column_number': col_num,
        'value': cell.value,
        'data_type': type(cell.value).__name__
    }

    print(f"\n{col_letter}{1} (Column {col_num}):")
    print(f"  Value: {cell.value}")
    print(f"  Type: {header_info['data_type']}")

    # Fill/Background
    if cell.fill and cell.fill.fgColor:
        rgb = str(cell.fill.fgColor.rgb)
        if rgb != '00000000':
            header_info['fill'] = {
                'pattern_type': cell.fill.patternType,
                'rgb': rgb,
                'hex': rgb_to_hex(rgb)
            }
            print(f"  Fill: {cell.fill.patternType} - RGB:{rgb} HEX:{rgb_to_hex(rgb)}")

    # Font
    if cell.font:
        font_details = []
        if cell.font.bold:
            font_details.append("Bold")
        if cell.font.italic:
            font_details.append("Italic")
        if cell.font.underline:
            font_details.append(f"Underline:{cell.font.underline}")
        if cell.font.size:
            font_details.append(f"Size:{cell.font.size}")
        if cell.font.name:
            font_details.append(f"Name:{cell.font.name}")
        if cell.font.color and str(cell.font.color.rgb) != '00000000':
            font_details.append(f"Color:{cell.font.color.rgb}")

        if font_details:
            header_info['font'] = font_details
            print(f"  Font: {', '.join(font_details)}")

    # Alignment
    if cell.alignment:
        align_details = []
        if cell.alignment.horizontal:
            align_details.append(f"H:{cell.alignment.horizontal}")
        if cell.alignment.vertical:
            align_details.append(f"V:{cell.alignment.vertical}")
        if cell.alignment.wrapText:
            align_details.append("Wrap:True")
        if cell.alignment.shrinkToFit:
            align_details.append("Shrink:True")
        if cell.alignment.textRotation:
            align_details.append(f"Rotation:{cell.alignment.textRotation}")

        if align_details:
            header_info['alignment'] = align_details
            print(f"  Alignment: {', '.join(align_details)}")

    # Borders
    if cell.border:
        border_details = []
        if cell.border.top and cell.border.top.style:
            border_details.append(f"Top:{cell.border.top.style}")
        if cell.border.bottom and cell.border.bottom.style:
            border_details.append(f"Bottom:{cell.border.bottom.style}")
        if cell.border.left and cell.border.left.style:
            border_details.append(f"Left:{cell.border.left.style}")
        if cell.border.right and cell.border.right.style:
            border_details.append(f"Right:{cell.border.right.style}")

        if border_details:
            header_info['border'] = border_details
            print(f"  Border: {', '.join(border_details)}")

    # Number format
    if cell.number_format and cell.number_format != 'General':
        header_info['number_format'] = cell.number_format
        print(f"  Number Format: {cell.number_format}")

    headers.append(header_info)

# 2. COLUMN INFORMATION
print(f"\n{'='*100}")
print("2. COLUMN WIDTHS AND PROPERTIES")
print(f"{'='*100}")

for col_num in range(1, 17):
    col_letter = get_column_letter(col_num)
    if col_letter in sheet.column_dimensions:
        col_dim = sheet.column_dimensions[col_letter]
        details = []
        if col_dim.width:
            details.append(f"Width:{col_dim.width:.2f}")
        if col_dim.hidden:
            details.append("Hidden:True")
        if col_dim.bestFit:
            details.append("BestFit:True")

        if details:
            print(f"  Column {col_letter}: {', '.join(details)}")

# 3. ROW HEIGHTS
print(f"\n{'='*100}")
print("3. ROW HEIGHTS (First 10 rows)")
print(f"{'='*100}")

for row_num in range(1, 11):
    if row_num in sheet.row_dimensions:
        row_dim = sheet.row_dimensions[row_num]
        details = []
        if row_dim.height:
            details.append(f"Height:{row_dim.height:.2f}")
        if row_dim.hidden:
            details.append("Hidden:True")

        if details:
            print(f"  Row {row_num}: {', '.join(details)}")

# 4. MERGED CELLS
print(f"\n{'='*100}")
print("4. MERGED CELLS")
print(f"{'='*100}")

if sheet.merged_cells:
    print(f"Total Merged Cell Ranges: {len(sheet.merged_cells.ranges)}")
    for merged_range in sheet.merged_cells.ranges:
        print(f"  {merged_range}")

# 5. DATA VALIDATIONS
print(f"\n{'='*100}")
print("5. DATA VALIDATIONS / DROPDOWNS")
print(f"{'='*100}")

if hasattr(sheet, 'data_validations') and sheet.data_validations:
    print(f"Total Validations: {len(sheet.data_validations.dataValidation)}")
    for i, dv in enumerate(sheet.data_validations.dataValidation, 1):
        print(f"\nValidation {i}:")
        print(f"  Type: {dv.type}")
        print(f"  Range: {dv.sqref}")
        if dv.formula1:
            print(f"  Formula1: {dv.formula1}")
        if dv.formula2:
            print(f"  Formula2: {dv.formula2}")
        if dv.allowBlank is not None:
            print(f"  Allow Blank: {dv.allowBlank}")
        if dv.showDropDown is not None:
            print(f"  Show Dropdown: {dv.showDropDown}")

# 6. CONDITIONAL FORMATTING
print(f"\n{'='*100}")
print("6. CONDITIONAL FORMATTING RULES")
print(f"{'='*100}")

if hasattr(sheet, 'conditional_formatting'):
    print(f"Total CF Ranges: {len(sheet.conditional_formatting._cf_rules)}")
    for cf_range, rules_list in sheet.conditional_formatting._cf_rules.items():
        print(f"\nRange: {cf_range}")
        for i, rule in enumerate(rules_list, 1):
            print(f"  Rule {i}:")
            if hasattr(rule, 'type'):
                print(f"    Type: {rule.type}")
            if hasattr(rule, 'formula') and rule.formula:
                print(f"    Formula: {rule.formula}")
            if hasattr(rule, 'text') and rule.text:
                print(f"    Text: {rule.text}")
            if hasattr(rule, 'dxf') and rule.dxf:
                if rule.dxf.fill and hasattr(rule.dxf.fill, 'bgColor'):
                    print(f"    Fill Color: {rule.dxf.fill.bgColor}")
                if rule.dxf.font:
                    if rule.dxf.font.bold:
                        print(f"    Font Bold: True")
                    if rule.dxf.font.color:
                        print(f"    Font Color: {rule.dxf.font.color}")

# 7. FORMULAS
print(f"\n{'='*100}")
print("7. FORMULAS (First 20 rows)")
print(f"{'='*100}")

formulas = []
for row in range(1, 21):
    for col in range(1, 17):
        cell = sheet.cell(row=row, column=col)
        if isinstance(cell.value, str) and cell.value.startswith('='):
            cell_ref = f"{get_column_letter(col)}{row}"
            formulas.append((cell_ref, cell.value))
            print(f"  {cell_ref}: {cell.value}")

if not formulas:
    print("  No formulas found in first 20 rows")

# 8. SAMPLE DATA (Rows 2-10)
print(f"\n{'='*100}")
print("8. SAMPLE DATA (Rows 2-10)")
print(f"{'='*100}")

for row_num in range(2, 11):
    print(f"\nRow {row_num}:")
    row_values = []
    for col_num in range(1, 17):
        cell = sheet.cell(row=row_num, column=col_num)
        if cell.value is not None:
            col_letter = get_column_letter(col_num)
            value_str = str(cell.value)[:50]  # Limit length
            row_values.append(f"{col_letter}={value_str}")

    if row_values:
        for val in row_values:
            print(f"  {val}")
    else:
        print("  (Empty row)")

# 9. CELL FORMATTING PATTERNS (First 10 rows)
print(f"\n{'='*100}")
print("9. UNIQUE CELL FORMATTING PATTERNS (Rows 1-10)")
print(f"{'='*100}")

unique_formats = {}
for row in range(1, 11):
    for col in range(1, 17):
        cell = sheet.cell(row=row, column=col)

        # Create format signature
        format_sig = []
        if cell.fill and cell.fill.fgColor:
            rgb = str(cell.fill.fgColor.rgb)
            if rgb != '00000000':
                format_sig.append(f"Fill:{rgb}")
        if cell.font and cell.font.bold:
            format_sig.append("Bold")
        if cell.alignment and cell.alignment.horizontal:
            format_sig.append(f"Align:{cell.alignment.horizontal}")

        if format_sig:
            sig_str = ', '.join(format_sig)
            if sig_str not in unique_formats:
                unique_formats[sig_str] = []
            unique_formats[sig_str].append(f"{get_column_letter(col)}{row}")

print("Unique Format Patterns Found:")
for i, (format_pattern, cells) in enumerate(unique_formats.items(), 1):
    print(f"\n  Pattern {i}: {format_pattern}")
    print(f"    Applied to cells: {', '.join(cells[:10])}")
    if len(cells) > 10:
        print(f"    ... and {len(cells) - 10} more")

# 10. PROTECTION
print(f"\n{'='*100}")
print("10. SHEET PROTECTION")
print(f"{'='*100}")

if hasattr(sheet, 'protection') and sheet.protection.sheet:
    print("Sheet Protection: ENABLED")
    if sheet.protection.password:
        print("  Password: SET")
    else:
        print("  Password: NONE")
else:
    print("Sheet Protection: DISABLED")

# 11. PRINT SETTINGS
print(f"\n{'='*100}")
print("11. PRINT SETTINGS")
print(f"{'='*100}")

if sheet.page_setup:
    print(f"  Orientation: {sheet.page_setup.orientation}")
    print(f"  Paper Size: {sheet.page_setup.paperSize}")
    if sheet.page_setup.fitToWidth:
        print(f"  Fit To Width: {sheet.page_setup.fitToWidth}")
    if sheet.page_setup.fitToHeight:
        print(f"  Fit To Height: {sheet.page_setup.fitToHeight}")

if sheet.print_options:
    print(f"  Grid Lines: {sheet.print_options.gridLines}")
    print(f"  Headings: {sheet.print_options.headings}")

# 12. SUMMARY FOR EXPORT
print(f"\n{'='*100}")
print("12. EXPORT REQUIREMENTS SUMMARY")
print(f"{'='*100}")

print("\nTo preserve ANEX-IV sheet formatting when exporting, you must:")
print("\n1. HEADER ROW (Row 1):")
print("   - Background: Light Blue (#C6D9F0)")
print("   - Font: Bold, Size 14, Calibri")
print("   - Alignment: Center horizontally and vertically")
print("   - Border: Medium borders on all sides")
print("   - Wrap text: Enabled")

print("\n2. COLUMN WIDTHS:")
for col_num in range(1, 6):
    col_letter = get_column_letter(col_num)
    if col_letter in sheet.column_dimensions and sheet.column_dimensions[col_letter].width:
        print(f"   - Column {col_letter}: {sheet.column_dimensions[col_letter].width:.2f}")

print("\n3. MERGED CELLS:")
print(f"   - Total: {len(sheet.merged_cells.ranges)} merged cell ranges")
print("   - Must preserve all merged ranges shown in section 4")

print("\n4. DATA VALIDATIONS:")
if hasattr(sheet, 'data_validations') and sheet.data_validations:
    print(f"   - Total: {len(sheet.data_validations.dataValidation)} validation rules")
    print("   - Mainly Y/N dropdowns on specific cells")

print("\n5. CONDITIONAL FORMATTING:")
if hasattr(sheet, 'conditional_formatting'):
    print(f"   - Total: {len(sheet.conditional_formatting._cf_rules)} CF ranges")
    print("   - Rules for highlighting blanks and data cells")

print("\n6. FORMULAS:")
print(f"   - Total found in first 20 rows: {len(formulas)}")
print("   - Must preserve formula references, not calculated values")

print(f"\n{'='*100}")
print("ANALYSIS COMPLETE")
print(f"{'='*100}")

wb.close()
