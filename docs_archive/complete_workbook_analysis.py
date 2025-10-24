import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Font, Border, Alignment, Protection
import json

def rgb_to_hex(rgb):
    """Convert RGB to hex color string"""
    if rgb and len(str(rgb)) >= 6:
        return f"#{str(rgb)[-6:]}"
    return None

def analyze_cell_formatting(cell):
    """Extract all formatting information from a cell"""
    formatting = {}

    # Fill/Background color
    if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
        rgb = str(cell.fill.fgColor.rgb)
        if rgb != '00000000':
            formatting['fill'] = {
                'pattern_type': cell.fill.patternType,
                'color': rgb,
                'hex': rgb_to_hex(rgb)
            }

    # Font
    if cell.font:
        font_info = {}
        if cell.font.bold:
            font_info['bold'] = True
        if cell.font.italic:
            font_info['italic'] = True
        if cell.font.underline:
            font_info['underline'] = cell.font.underline
        if cell.font.size:
            font_info['size'] = cell.font.size
        if cell.font.name:
            font_info['name'] = cell.font.name
        if cell.font.color and cell.font.color.rgb and str(cell.font.color.rgb) != '00000000':
            font_info['color'] = str(cell.font.color.rgb)
        if font_info:
            formatting['font'] = font_info

    # Alignment
    if cell.alignment:
        alignment_info = {}
        if cell.alignment.horizontal:
            alignment_info['horizontal'] = cell.alignment.horizontal
        if cell.alignment.vertical:
            alignment_info['vertical'] = cell.alignment.vertical
        if cell.alignment.wrapText:
            alignment_info['wrap_text'] = True
        if cell.alignment.shrinkToFit:
            alignment_info['shrink_to_fit'] = True
        if cell.alignment.textRotation:
            alignment_info['text_rotation'] = cell.alignment.textRotation
        if alignment_info:
            formatting['alignment'] = alignment_info

    # Borders
    if cell.border:
        border_info = {}
        if cell.border.top and cell.border.top.style:
            border_info['top'] = cell.border.top.style
        if cell.border.bottom and cell.border.bottom.style:
            border_info['bottom'] = cell.border.bottom.style
        if cell.border.left and cell.border.left.style:
            border_info['left'] = cell.border.left.style
        if cell.border.right and cell.border.right.style:
            border_info['right'] = cell.border.right.style
        if border_info:
            formatting['border'] = border_info

    # Number format
    if cell.number_format and cell.number_format != 'General':
        formatting['number_format'] = cell.number_format

    # Protection
    if cell.protection:
        protection_info = {}
        if hasattr(cell.protection, 'locked'):
            protection_info['locked'] = cell.protection.locked
        if hasattr(cell.protection, 'hidden'):
            protection_info['hidden'] = cell.protection.hidden
        if protection_info:
            formatting['protection'] = protection_info

    return formatting if formatting else None

def analyze_data_validation(sheet):
    """Extract data validation rules"""
    validations = []
    if hasattr(sheet, 'data_validations') and sheet.data_validations:
        for dv in sheet.data_validations.dataValidation:
            validation_info = {
                'type': dv.type,
                'sqref': str(dv.sqref),
            }
            if dv.formula1:
                validation_info['formula1'] = dv.formula1
            if dv.formula2:
                validation_info['formula2'] = dv.formula2
            if dv.allowBlank is not None:
                validation_info['allow_blank'] = dv.allowBlank
            if dv.showDropDown is not None:
                validation_info['show_dropdown'] = dv.showDropDown
            if dv.showErrorMessage:
                validation_info['show_error_message'] = True
                if dv.error:
                    validation_info['error_message'] = dv.error
            if dv.showInputMessage:
                validation_info['show_input_message'] = True
                if dv.prompt:
                    validation_info['input_message'] = dv.prompt
            validations.append(validation_info)
    return validations

def analyze_conditional_formatting(sheet):
    """Extract conditional formatting rules"""
    cf_rules = []
    if hasattr(sheet, 'conditional_formatting') and sheet.conditional_formatting:
        for cf_range, rules_list in sheet.conditional_formatting._cf_rules.items():
            for rule in rules_list:
                rule_info = {
                    'range': str(cf_range),
                    'type': rule.type if hasattr(rule, 'type') else None,
                }
                if hasattr(rule, 'formula') and rule.formula:
                    rule_info['formula'] = rule.formula
                if hasattr(rule, 'dxf') and rule.dxf:
                    dxf_info = {}
                    if rule.dxf.fill and rule.dxf.fill.bgColor:
                        dxf_info['fill_color'] = str(rule.dxf.fill.bgColor.rgb) if hasattr(rule.dxf.fill.bgColor, 'rgb') else None
                    if rule.dxf.font:
                        font_info = {}
                        if rule.dxf.font.bold:
                            font_info['bold'] = True
                        if rule.dxf.font.color:
                            font_info['color'] = str(rule.dxf.font.color.rgb) if hasattr(rule.dxf.font.color, 'rgb') else None
                        if font_info:
                            dxf_info['font'] = font_info
                    if dxf_info:
                        rule_info['formatting'] = dxf_info
                cf_rules.append(rule_info)
    return cf_rules

def get_used_range(sheet):
    """Get the actual used range of data in a sheet"""
    max_row = sheet.max_row
    max_col = sheet.max_column

    # Find actual last row with data
    actual_max_row = 0
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = sheet.cell(row=row, column=col)
            if cell.value is not None or (cell.fill and cell.fill.fgColor and str(cell.fill.fgColor.rgb) != '00000000'):
                actual_max_row = max(actual_max_row, row)

    # Find actual last column with data
    actual_max_col = 0
    for col in range(1, max_col + 1):
        for row in range(1, max_row + 1):
            cell = sheet.cell(row=row, column=col)
            if cell.value is not None or (cell.fill and cell.fill.fgColor and str(cell.fill.fgColor.rgb) != '00000000'):
                actual_max_col = max(actual_max_col, col)

    return actual_max_row, actual_max_col

def analyze_sheet(sheet, sheet_name):
    """Comprehensive analysis of a single sheet"""
    print(f"\n{'='*100}")
    print(f"SHEET: {sheet_name}")
    print(f"{'='*100}")

    sheet_info = {
        'name': sheet_name,
        'max_row': sheet.max_row,
        'max_column': sheet.max_column,
    }

    # Get actual used range
    actual_max_row, actual_max_col = get_used_range(sheet)
    sheet_info['actual_max_row'] = actual_max_row
    sheet_info['actual_max_column'] = actual_max_col
    sheet_info['actual_max_column_letter'] = get_column_letter(actual_max_col) if actual_max_col > 0 else 'A'

    print(f"\nDimensions:")
    print(f"  Max Row: {sheet.max_row}")
    print(f"  Max Column: {sheet.max_column} ({get_column_letter(sheet.max_column)})")
    print(f"  Actual Used Row: {actual_max_row}")
    print(f"  Actual Used Column: {actual_max_col} ({get_column_letter(actual_max_col) if actual_max_col > 0 else 'A'})")

    # Check for merged cells
    merged_cells = []
    if sheet.merged_cells:
        print(f"\nMerged Cells: {len(sheet.merged_cells.ranges)}")
        for merged_range in sheet.merged_cells.ranges:
            merged_cells.append(str(merged_range))
            print(f"  {merged_range}")
    sheet_info['merged_cells'] = merged_cells

    # Check for hidden rows
    hidden_rows = []
    for row_num in range(1, actual_max_row + 1):
        if row_num in sheet.row_dimensions and sheet.row_dimensions[row_num].hidden:
            hidden_rows.append(row_num)
    if hidden_rows:
        print(f"\nHidden Rows: {hidden_rows}")
    sheet_info['hidden_rows'] = hidden_rows

    # Check for hidden columns
    hidden_cols = []
    for col_num in range(1, actual_max_col + 1):
        col_letter = get_column_letter(col_num)
        if col_letter in sheet.column_dimensions and sheet.column_dimensions[col_letter].hidden:
            hidden_cols.append(col_letter)
    if hidden_cols:
        print(f"\nHidden Columns: {hidden_cols}")
    sheet_info['hidden_columns'] = hidden_cols

    # Column widths
    print(f"\nColumn Widths (non-default):")
    column_widths = {}
    for col_num in range(1, actual_max_col + 1):
        col_letter = get_column_letter(col_num)
        if col_letter in sheet.column_dimensions:
            width = sheet.column_dimensions[col_letter].width
            if width and width != 8.43:  # Default width
                column_widths[col_letter] = width
                print(f"  Column {col_letter}: {width:.2f}")
    sheet_info['column_widths'] = column_widths

    # Row heights
    print(f"\nRow Heights (non-default):")
    row_heights = {}
    for row_num in range(1, min(actual_max_row + 1, 20)):  # First 20 rows
        if row_num in sheet.row_dimensions:
            height = sheet.row_dimensions[row_num].height
            if height and height != 15:  # Default height
                row_heights[row_num] = height
                print(f"  Row {row_num}: {height:.2f}")
    sheet_info['row_heights'] = row_heights

    # Data validations
    validations = analyze_data_validation(sheet)
    if validations:
        print(f"\nData Validations: {len(validations)}")
        for i, validation in enumerate(validations[:5]):  # Show first 5
            print(f"  {i+1}. Range: {validation['sqref']}, Type: {validation['type']}")
            if 'formula1' in validation:
                print(f"     Formula: {validation['formula1']}")
    sheet_info['data_validations'] = validations

    # Conditional formatting
    cf_rules = analyze_conditional_formatting(sheet)
    if cf_rules:
        print(f"\nConditional Formatting Rules: {len(cf_rules)}")
        for i, rule in enumerate(cf_rules[:5]):  # Show first 5
            print(f"  {i+1}. Range: {rule['range']}, Type: {rule.get('type', 'N/A')}")
    sheet_info['conditional_formatting'] = cf_rules

    # Sheet protection
    protection_info = {}
    if hasattr(sheet, 'protection') and sheet.protection.sheet:
        print(f"\nSheet Protection: ENABLED")
        protection_info['enabled'] = True
        protection_info['password'] = 'Set' if sheet.protection.password else 'None'
        print(f"  Password: {protection_info['password']}")
    else:
        print(f"\nSheet Protection: DISABLED")
        protection_info['enabled'] = False
    sheet_info['protection'] = protection_info

    # Formulas
    print(f"\nFormulas (sample):")
    formulas = []
    formula_count = 0
    for row in range(1, min(actual_max_row + 1, 50)):  # Check first 50 rows
        for col in range(1, actual_max_col + 1):
            cell = sheet.cell(row=row, column=col)
            if isinstance(cell.value, str) and cell.value.startswith('='):
                formula_count += 1
                cell_ref = f"{get_column_letter(col)}{row}"
                formulas.append({'cell': cell_ref, 'formula': cell.value})
                if formula_count <= 10:  # Show first 10
                    print(f"  {cell_ref}: {cell.value}")
    print(f"  Total formulas found (first 50 rows): {formula_count}")
    sheet_info['formulas'] = formulas

    return sheet_info

def analyze_annex4_headers(sheet):
    """Detailed analysis of Annex 4 headers"""
    print(f"\n{'='*100}")
    print(f"DETAILED ANALYSIS: ANNEX 4 HEADERS (Row 1)")
    print(f"{'='*100}")

    headers = []
    max_col = sheet.max_column

    print(f"\nColumn Headers:")
    for col_num in range(1, max_col + 1):
        cell = sheet.cell(row=1, column=col_num)
        col_letter = get_column_letter(col_num)

        if cell.value:
            header_info = {
                'column': col_letter,
                'column_number': col_num,
                'value': cell.value,
                'data_type': str(type(cell.value).__name__)
            }

            # Get formatting
            formatting = analyze_cell_formatting(cell)
            if formatting:
                header_info['formatting'] = formatting

            headers.append(header_info)

            print(f"\n  {col_letter} (Column {col_num}): '{cell.value}'")
            print(f"    Data Type: {header_info['data_type']}")

            if formatting:
                if 'fill' in formatting:
                    print(f"    Fill: {formatting['fill']['color']}")
                if 'font' in formatting:
                    font_details = []
                    if formatting['font'].get('bold'):
                        font_details.append('Bold')
                    if formatting['font'].get('size'):
                        font_details.append(f"Size={formatting['font']['size']}")
                    if formatting['font'].get('color'):
                        font_details.append(f"Color={formatting['font']['color']}")
                    if font_details:
                        print(f"    Font: {', '.join(font_details)}")
                if 'alignment' in formatting:
                    align_details = []
                    if formatting['alignment'].get('horizontal'):
                        align_details.append(f"H={formatting['alignment']['horizontal']}")
                    if formatting['alignment'].get('vertical'):
                        align_details.append(f"V={formatting['alignment']['vertical']}")
                    if formatting['alignment'].get('wrap_text'):
                        align_details.append('Wrap=True')
                    if align_details:
                        print(f"    Alignment: {', '.join(align_details)}")
                if 'border' in formatting:
                    print(f"    Border: {formatting['border']}")

    # Check for data in rows 2-10
    print(f"\n\nData Check (Rows 2-10):")
    for row_num in range(2, 11):
        row_data = []
        for col_num in range(1, max_col + 1):
            cell = sheet.cell(row=row_num, column=col_num)
            if cell.value:
                row_data.append(f"{get_column_letter(col_num)}={cell.value}")
        if row_data:
            print(f"  Row {row_num}: {', '.join(row_data[:5])}")  # Show first 5 values
        else:
            print(f"  Row {row_num}: (empty)")

    return headers

# Main analysis
print(f"{'='*100}")
print(f"COMPREHENSIVE EXCEL WORKBOOK ANALYSIS")
print(f"File: QAQC-FS8-Acceptance.xlsx")
print(f"{'='*100}")

# Load workbook
wb = openpyxl.load_workbook('/Users/sunildeesu/myProjects/Hemanth/DataEntryWeb/QAQC-FS8-Acceptance.xlsx', data_only=False)

# List all sheets
print(f"\n{'='*100}")
print(f"WORKBOOK SHEETS")
print(f"{'='*100}")
print(f"\nTotal Sheets: {len(wb.sheetnames)}")
print(f"\nSheet Names (in order):")
for i, sheet_name in enumerate(wb.sheetnames, 1):
    print(f"  {i}. {sheet_name}")

# Analyze each sheet
all_sheets_info = {}
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    sheet_info = analyze_sheet(sheet, sheet_name)
    all_sheets_info[sheet_name] = sheet_info

# Special analysis for Annex 4
if 'Annex 4' in wb.sheetnames:
    annex4_sheet = wb['Annex 4']
    annex4_headers = analyze_annex4_headers(annex4_sheet)
    all_sheets_info['Annex 4']['headers_detailed'] = annex4_headers
elif 'ANEX-IV' in wb.sheetnames:
    annex4_sheet = wb['ANEX-IV']
    annex4_headers = analyze_annex4_headers(annex4_sheet)
    all_sheets_info['ANEX-IV']['headers_detailed'] = annex4_headers

# Save complete analysis to JSON
output_file = '/Users/sunildeesu/myProjects/Hemanth/DataEntryWeb/workbook_analysis.json'
with open(output_file, 'w', encoding='utf-8') as f:
    json.dump(all_sheets_info, f, indent=2, ensure_ascii=False)

print(f"\n{'='*100}")
print(f"ANALYSIS COMPLETE")
print(f"{'='*100}")
print(f"\nDetailed analysis saved to: {output_file}")

wb.close()
