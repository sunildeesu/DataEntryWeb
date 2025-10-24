import openpyxl
import pandas as pd

# Load the workbook
wb = openpyxl.load_workbook('QAQC-FS8-Acceptance.xlsx')

print("="*100)
print("EXTRACTING ALL COLUMN HEADERS FROM ENTRY SHEET")
print("="*100)

# Get ENTRY sheet
entry_sheet = wb['ENTRY']

# Read headers from row 10 (the main header row for village data)
print("\nRow 10 Headers (Main village data columns):")
print("-"*100)

headers_row10 = []
for col_idx, col in enumerate(entry_sheet[10], 1):
    if col.value:
        headers_row10.append({
            'col': col_idx,
            'letter': openpyxl.utils.get_column_letter(col_idx),
            'value': str(col.value).strip()
        })
        print(f"Column {col_idx} ({openpyxl.utils.get_column_letter(col_idx)}): {col.value}")

# Also check row 9 for any merged headers or sub-headers
print("\n\nRow 9 Headers (Sub-headers or category headers):")
print("-"*100)

headers_row9 = []
for col_idx, col in enumerate(entry_sheet[9], 1):
    if col.value:
        headers_row9.append({
            'col': col_idx,
            'letter': openpyxl.utils.get_column_letter(col_idx),
            'value': str(col.value).strip()
        })
        print(f"Column {col_idx} ({openpyxl.utils.get_column_letter(col_idx)}): {col.value}")

# Check for merged cells in header area
print("\n\nMerged Cells in Header Area (rows 9-10):")
print("-"*100)

for merged_range in entry_sheet.merged_cells.ranges:
    if merged_range.min_row in [9, 10]:
        # Get the value from the merged cell
        cell_coord = f"{openpyxl.utils.get_column_letter(merged_range.min_col)}{merged_range.min_row}"
        cell_value = entry_sheet[cell_coord].value
        print(f"{merged_range} → '{cell_value}'")

# Count total columns with data
print("\n\nColumn Statistics:")
print("-"*100)
print(f"Total columns with headers in row 10: {len(headers_row10)}")
print(f"Total columns with headers in row 9: {len(headers_row9)}")

# Check the specific columns mentioned by user
print("\n\nChecking Specific Columns (File Naming, GPS, etc.):")
print("-"*100)

# Search for keywords in headers
keywords = [
    "File naming", "File Naming",
    "Raw & RINEX", "RINEX",
    "GPS Log", "Log Sheet",
    "Base line", "Baseline",
    "Network adjustment", "Network Adjustment",
    "UAV processing", "UAV Processing",
    "PPK", "Flight Log",
    "Projection", "Datum",
    "QA/QC Report", "QAQC"
]

for keyword in keywords:
    for header in headers_row10:
        if keyword.lower() in header['value'].lower():
            print(f"✓ Found '{keyword}' in Column {header['letter']}: {header['value']}")
    for header in headers_row9:
        if keyword.lower() in header['value'].lower():
            print(f"✓ Found '{keyword}' in Row 9, Column {header['letter']}: {header['value']}")

# Export complete header structure
print("\n\n" + "="*100)
print("COMPLETE HEADER MAPPING (for HTML table)")
print("="*100)

# Combine row 9 and row 10 to understand the full structure
for col_idx in range(1, 50):  # Check first 50 columns
    col_letter = openpyxl.utils.get_column_letter(col_idx)
    row9_val = entry_sheet[f'{col_letter}9'].value
    row10_val = entry_sheet[f'{col_letter}10'].value

    if row9_val or row10_val:
        print(f"\nColumn {col_letter} (#{col_idx}):")
        if row9_val:
            print(f"  Row 9:  {row9_val}")
        if row10_val:
            print(f"  Row 10: {row10_val}")

wb.close()
