import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Alignment

# Load the workbook
wb = openpyxl.load_workbook('QAQC-FS8-Acceptance.xlsx')
entry_sheet = wb['ENTRY']

print("="*100)
print("COMPREHENSIVE EXCEL FORMATTING ANALYSIS - ENTRY SHEET")
print("="*100)

# 1. Check metadata cells (rows 1-7) for special formatting
print("\n1. METADATA SECTION FORMATTING (Rows 1-7)")
print("-"*100)

metadata_cells = [
    ('E1', 'Examiner Name'),
    ('E2', 'Name of Examiner'),
    ('E3', 'Designation'),
    ('E5', 'Commenced On'),
    ('E6', 'Finished On'),
    ('H5', 'District'),
    ('H6', 'Taluk'),
    ('J2', 'Data Received'),
    ('J3', 'No of Days'),
    ('F7', 'Agency'),
    ('J6', 'Submission Type'),
    ('L3', 'Hard Disk No'),
]

for cell_ref, desc in metadata_cells:
    cell = entry_sheet[cell_ref]
    print(f"\n{cell_ref} ({desc}):")
    print(f"  Value: {cell.value}")

    if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb != '00000000':
        print(f"  Fill: {cell.fill.patternType} - {cell.fill.fgColor.rgb}")

    if cell.font:
        if cell.font.bold:
            print(f"  Font: Bold")
        if cell.font.color and cell.font.color.rgb != '00000000':
            print(f"  Font Color: {cell.font.color.rgb}")
        if cell.font.size:
            print(f"  Font Size: {cell.font.size}")

    if cell.alignment:
        if cell.alignment.horizontal:
            print(f"  Alignment: {cell.alignment.horizontal}")
        if cell.alignment.vertical:
            print(f"  Vertical: {cell.alignment.vertical}")

# 2. Check header rows (rows 9-10)
print("\n\n2. TABLE HEADER FORMATTING (Rows 9-10)")
print("-"*100)

sample_header_cells = ['A9', 'C9', 'D9', 'E9', 'M9', 'N9', 'O9', 'AG9', 'AH9', 'AI9', 'AJ9']

for cell_ref in sample_header_cells:
    cell = entry_sheet[cell_ref]
    print(f"\n{cell_ref}:")
    if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb != '00000000':
        print(f"  Fill: {cell.fill.fgColor.rgb}")
    if cell.font:
        if cell.font.bold:
            print(f"  Font: Bold")
        if cell.font.color and cell.font.color.rgb != '00000000':
            print(f"  Font Color: {cell.font.color.rgb}")
    if cell.alignment:
        print(f"  Alignment: H={cell.alignment.horizontal}, V={cell.alignment.vertical}")
        if cell.alignment.wrapText:
            print(f"  Wrap Text: Yes")

# 3. Check data row formatting (row 11 as sample)
print("\n\n3. DATA ROW FORMATTING (Row 11 - Sample)")
print("-"*100)

# Check specific columns
data_cols = ['A', 'C', 'D', 'E', 'F', 'G', 'H', 'M', 'N', 'O', 'S', 'V', 'Z', 'AC', 'AG', 'AH', 'AI', 'AJ', 'AK']

for col in data_cols:
    cell = entry_sheet[f'{col}11']
    if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb != '00000000':
        print(f"\n{col}11: Fill={cell.fill.fgColor.rgb}")
        if cell.value:
            print(f"  Value: {cell.value}")

# 4. Check for cell borders
print("\n\n4. BORDER STYLES")
print("-"*100)

sample_cell = entry_sheet['A9']
if sample_cell.border:
    print(f"Sample header cell (A9) border:")
    print(f"  Top: {sample_cell.border.top.style if sample_cell.border.top else 'None'}")
    print(f"  Bottom: {sample_cell.border.bottom.style if sample_cell.border.bottom else 'None'}")
    print(f"  Left: {sample_cell.border.left.style if sample_cell.border.left else 'None'}")
    print(f"  Right: {sample_cell.border.right.style if sample_cell.border.right else 'None'}")

sample_cell = entry_sheet['A11']
if sample_cell.border:
    print(f"\nSample data cell (A11) border:")
    print(f"  Top: {sample_cell.border.top.style if sample_cell.border.top else 'None'}")
    print(f"  Bottom: {sample_cell.border.bottom.style if sample_cell.border.bottom else 'None'}")
    print(f"  Left: {sample_cell.border.left.style if sample_cell.border.left else 'None'}")
    print(f"  Right: {sample_cell.border.right.style if sample_cell.border.right else 'None'}")

# 5. Check column widths
print("\n\n5. COLUMN WIDTHS")
print("-"*100)

important_cols = ['A', 'C', 'D', 'E', 'M', 'N', 'O', 'S', 'T', 'U', 'V', 'AG', 'AH', 'AI', 'AJ', 'AK']
for col in important_cols:
    col_letter = col
    if col_letter in entry_sheet.column_dimensions:
        width = entry_sheet.column_dimensions[col_letter].width
        if width:
            print(f"Column {col_letter}: {width:.2f}")

# 6. Check row heights
print("\n\n6. ROW HEIGHTS")
print("-"*100)

for row_num in [9, 10, 11, 12]:
    if row_num in entry_sheet.row_dimensions:
        height = entry_sheet.row_dimensions[row_num].height
        if height:
            print(f"Row {row_num}: {height:.2f}")

# 7. Check for number formatting
print("\n\n7. NUMBER FORMATTING")
print("-"*100)

number_cells = ['S11', 'T11', 'U11', 'V11', 'Z11', 'AG11', 'AH11']
for cell_ref in number_cells:
    cell = entry_sheet[cell_ref]
    if cell.number_format and cell.number_format != 'General':
        print(f"{cell_ref}: {cell.number_format}")
        print(f"  Value: {cell.value}")

# 8. Check for protection
print("\n\n8. SHEET PROTECTION")
print("-"*100)

if entry_sheet.protection.sheet:
    print("Sheet is protected")
    print(f"  Password: {'Set' if entry_sheet.protection.password else 'None'}")
else:
    print("Sheet is not protected")

# 9. Summary of all unique fill colors used
print("\n\n9. ALL UNIQUE FILL COLORS IN DATA AREA (Rows 11-17)")
print("-"*100)

unique_colors = set()
for row in range(11, 18):
    for col in range(1, 42):  # Columns A to AO
        cell = entry_sheet.cell(row=row, column=col)
        if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb and cell.fill.fgColor.rgb != '00000000':
            unique_colors.add((openpyxl.utils.get_column_letter(col), cell.fill.fgColor.rgb))

print("Unique background colors found:")
for col_letter, color in sorted(unique_colors):
    # Check if it's from conditional formatting or cell formatting
    print(f"  Column {col_letter}: {color}")

# 10. Check text alignment in data cells
print("\n\n10. TEXT ALIGNMENT IN DATA CELLS")
print("-"*100)

sample_data_cells = ['A11', 'D11', 'E11', 'M11', 'N11', 'S11', 'AI11', 'AK11']
for cell_ref in sample_data_cells:
    cell = entry_sheet[cell_ref]
    if cell.alignment:
        h_align = cell.alignment.horizontal or 'general'
        v_align = cell.alignment.vertical or 'bottom'
        wrap = cell.alignment.wrapText or False
        print(f"{cell_ref}: H={h_align}, V={v_align}, Wrap={wrap}")

wb.close()

print("\n\n" + "="*100)
print("ANALYSIS COMPLETE")
print("="*100)
