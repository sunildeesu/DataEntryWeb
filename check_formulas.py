import openpyxl
import pandas as pd

# Load the workbook
wb = openpyxl.load_workbook('QAQC-FS8-Acceptance.xlsx', data_only=False)

print("="*100)
print("CHECKING ALL SHEETS FOR FORMULAS")
print("="*100)

total_formulas = 0
sheets_with_formulas = []

for sheet_name in wb.sheetnames:
    print(f"\n{'='*100}")
    print(f"Sheet: {sheet_name}")
    print(f"{'='*100}")

    ws = wb[sheet_name]
    formulas_in_sheet = []

    # Check each cell for formulas
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formulas_in_sheet.append({
                    'cell': cell.coordinate,
                    'formula': cell.value
                })

    if formulas_in_sheet:
        sheets_with_formulas.append(sheet_name)
        total_formulas += len(formulas_in_sheet)
        print(f"Found {len(formulas_in_sheet)} formulas:")

        # Show first 20 formulas
        for formula_info in formulas_in_sheet[:20]:
            print(f"  {formula_info['cell']}: {formula_info['formula']}")

        if len(formulas_in_sheet) > 20:
            print(f"  ... and {len(formulas_in_sheet) - 20} more formulas")
    else:
        print("No formulas found in this sheet")

print("\n" + "="*100)
print("SUMMARY")
print("="*100)
print(f"Total sheets: {len(wb.sheetnames)}")
print(f"Sheets with formulas: {len(sheets_with_formulas)}")
print(f"Total formulas: {total_formulas}")

if sheets_with_formulas:
    print(f"\nSheets containing formulas:")
    for sheet in sheets_with_formulas:
        print(f"  - {sheet}")

# Check for data validation
print("\n" + "="*100)
print("CHECKING FOR DATA VALIDATION RULES")
print("="*100)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    if ws.data_validations.dataValidation:
        print(f"\nSheet '{sheet_name}' has data validation rules:")
        for dv in ws.data_validations.dataValidation:
            print(f"  Range: {dv.sqref}")
            print(f"  Type: {dv.type}")
            if dv.formula1:
                print(f"  Formula1: {dv.formula1}")
            if dv.formula2:
                print(f"  Formula2: {dv.formula2}")

# Check for named ranges
print("\n" + "="*100)
print("CHECKING FOR NAMED RANGES")
print("="*100)

if wb.defined_names:
    print(f"Found {len(wb.defined_names.definedName)} named ranges:")
    for name in wb.defined_names.definedName:
        print(f"  {name.name}: {name.value}")
else:
    print("No named ranges found")

# Check for conditional formatting
print("\n" + "="*100)
print("CHECKING FOR CONDITIONAL FORMATTING")
print("="*100)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    if ws.conditional_formatting:
        print(f"\nSheet '{sheet_name}' has conditional formatting")
        # Conditional formatting is complex, just note its presence

# Check for merged cells
print("\n" + "="*100)
print("CHECKING FOR MERGED CELLS")
print("="*100)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    if ws.merged_cells:
        print(f"\nSheet '{sheet_name}' has {len(ws.merged_cells.ranges)} merged cell ranges:")
        for merged_range in list(ws.merged_cells.ranges)[:10]:
            print(f"  {merged_range}")
        if len(ws.merged_cells.ranges) > 10:
            print(f"  ... and {len(ws.merged_cells.ranges) - 10} more")

wb.close()

print("\n" + "="*100)
print("ANALYSIS COMPLETE")
print("="*100)
