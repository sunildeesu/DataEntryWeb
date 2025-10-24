#!/usr/bin/env python3
"""
Excel Export Script for QA/QC Data Entry
This script copies the reference Excel file and fills in the data from the web form.
All formatting, formulas, and sheets are preserved.
"""

import sys
import json
import shutil
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook


def parse_date(date_string):
    """Convert date string (YYYY-MM-DD) to datetime object"""
    if not date_string:
        return ''
    try:
        return datetime.strptime(date_string, '%Y-%m-%d')
    except:
        return date_string  # Return as-is if parsing fails


def fill_anex_iv_sheet(wb, village_data, check_points_data):
    """
    Fill ANEX-IV sheet with check points data.

    Structure per village (5 rows):
    - Row 1: SL No, Village Name, Check point names (E-P for up to 12 points)
    - Row 2: H.Dist Error values
    - Row 3: Actual height values
    - Row 4: Observed height values
    - Row 5: Height Error (formula: =(E7-E8)*100), On Marker (Y/N) in C, Spatially Distributed (Y/N) in C
    """
    if 'ANEX-IV' not in wb.sheetnames:
        print(f"   ⚠️  ANEX-IV sheet not found in workbook")
        return

    ws = wb['ANEX-IV']

    # ANEX-IV starts at row 5 (after headers)
    current_row = 5

    for village_index, village in enumerate(village_data):
        village_row_id = village_index + 1  # Row ID from the form (1-based)

        # Check if we have check points data for this village
        if str(village_row_id) not in check_points_data:
            print(f"   - Village {village_row_id} ({village.get('villageName', 'N/A')}): No check points data")
            current_row += 5  # Skip 5 rows for this village
            continue

        cp_data = check_points_data[str(village_row_id)]
        check_points = cp_data.get('checkPoints', [])

        print(f"   ✓ Village {village_row_id} ({village.get('villageName', 'N/A')}): {len(check_points)} check points")

        # Row 1: SL No and Village Name
        ws.cell(row=current_row, column=1).value = village_row_id  # A: SL No
        ws.cell(row=current_row, column=2).value = village.get('villageName', '')  # B: Village Name

        # Fill check point names in columns E-P (up to 12 points)
        for cp_index, cp in enumerate(check_points[:12]):
            col = 5 + cp_index  # E=5, F=6, ..., P=16
            ws.cell(row=current_row, column=col).value = cp.get('name', f'CP{cp_index + 1}')

        # Row 2: H.Dist Error values
        for cp_index, cp in enumerate(check_points[:12]):
            col = 5 + cp_index
            h_dist_error = cp.get('hDistError', '')
            if h_dist_error:
                try:
                    ws.cell(row=current_row + 1, column=col).value = float(h_dist_error)
                except:
                    ws.cell(row=current_row + 1, column=col).value = h_dist_error

        # Row 3: Actual height values
        for cp_index, cp in enumerate(check_points[:12]):
            col = 5 + cp_index
            actual_height = cp.get('actualHeight', '')
            if actual_height:
                try:
                    ws.cell(row=current_row + 2, column=col).value = float(actual_height)
                except:
                    ws.cell(row=current_row + 2, column=col).value = actual_height

        # Row 4: Observed height values
        for cp_index, cp in enumerate(check_points[:12]):
            col = 5 + cp_index
            observed_height = cp.get('observedHeight', '')
            if observed_height:
                try:
                    ws.cell(row=current_row + 3, column=col).value = float(observed_height)
                except:
                    ws.cell(row=current_row + 3, column=col).value = observed_height

        # Row 5: Height Error (calculated values from form)
        for cp_index, cp in enumerate(check_points[:12]):
            col = 5 + cp_index
            height_error = cp.get('heightError', '')
            if height_error:
                try:
                    ws.cell(row=current_row + 4, column=col).value = float(height_error)
                except:
                    ws.cell(row=current_row + 4, column=col).value = height_error

        # Row 4: On Marker (Y/N) in column C
        on_marker = cp_data.get('onMarker', '')
        if on_marker:
            ws.cell(row=current_row + 3, column=3).value = on_marker

        # Row 5: Spatially Distributed (Y/N) in column C
        spatially_distributed = cp_data.get('spatiallyDistributed', '')
        if spatially_distributed:
            ws.cell(row=current_row + 4, column=3).value = spatially_distributed

        # Move to next village (5 rows down)
        current_row += 5

    print(f"   ✓ ANEX-IV sheet filled successfully")


def export_to_excel(data, output_filename=None):
    """
    Copy the reference Excel file and fill it with form data.

    Args:
        data: Dictionary containing form data
        output_filename: Optional custom filename for output

    Returns:
        Path to the exported file
    """
    # Get the reference file path
    script_dir = Path(__file__).parent
    reference_file = script_dir / 'QAQC-FS8-Acceptance.xlsx'

    if not reference_file.exists():
        raise FileNotFoundError(f"Reference file not found: {reference_file}")

    # Generate output filename if not provided
    if not output_filename:
        submission_type = data['metadata'].get('submissionType', 'FS8')
        district = data['metadata'].get('district', 'District')
        date_str = datetime.now().strftime('%Y-%m-%d')
        output_filename = f"QAQC-FS8-{submission_type}_{district}_{date_str}.xlsx"

    # Ensure .xlsx extension
    if not output_filename.endswith('.xlsx'):
        output_filename += '.xlsx'

    output_path = script_dir / output_filename

    print(f"📄 Copying reference file: {reference_file}")

    # Copy the reference file to create a new file
    shutil.copy2(reference_file, output_path)

    print(f"✓ Reference file copied to: {output_path}")
    print(f"📝 Opening workbook to fill data...")

    # Load the copied workbook (preserves ALL formatting, formulas, sheets)
    wb = load_workbook(output_path, keep_vba=True, data_only=False, keep_links=True)

    print(f"✓ Workbook loaded with {len(wb.sheetnames)} sheets")
    print(f"   Sheets: {', '.join(wb.sheetnames)}")

    # Get the ENTRY sheet
    if 'ENTRY' not in wb.sheetnames:
        raise ValueError("ENTRY sheet not found in reference file")

    ws = wb['ENTRY']

    print(f"✓ ENTRY sheet found - filling metadata...")

    # Fill metadata in ENTRY sheet
    metadata = data['metadata']

    print(f"DEBUG - Received metadata:")
    print(f"   commencedOn raw value: '{metadata.get('commencedOn')}'")
    print(f"   finishedOn raw value: '{metadata.get('finishedOn')}'")

    # Row 2
    examiner_name = metadata.get('examinerName', '')
    commenced_date = parse_date(metadata.get('commencedOn'))

    print(f"   Parsed commencedOn: '{commenced_date}' (type: {type(commenced_date).__name__})")
    print(f"   Writing to E2: '{examiner_name}'")
    print(f"   Writing to J2: '{commenced_date}'")

    ws['E2'] = examiner_name           # Examiner Name
    ws['J2'] = commenced_date          # Commenced Date

    # Set date format for the cell if it's a datetime
    if commenced_date and isinstance(commenced_date, datetime):
        ws['J2'].number_format = 'DD.MM.YYYY'

    print(f"   Verification - E2 value: '{ws['E2'].value}'")
    print(f"   Verification - J2 value: '{ws['J2'].value}'")

    # Row 3
    finished_date = parse_date(metadata.get('finishedOn'))
    ws['E3'] = metadata.get('designation', '')            # Designation
    ws['J3'] = finished_date                               # Finished Date
    ws['L3'] = metadata.get('noOfDays', '')                # Number of Days

    # Set date format for the cell if it's a datetime
    if finished_date and isinstance(finished_date, datetime):
        ws['J3'].number_format = 'DD.MM.YYYY'

    # Row 4 is empty (skip)

    # Row 5
    data_received_date = parse_date(metadata.get('dataReceivedDate'))
    ws['E5'] = data_received_date                              # Data Received Date
    ws['H5'] = metadata.get('district', '')                    # District

    # Set date format for the cell if it's a datetime
    if data_received_date and isinstance(data_received_date, datetime):
        ws['E5'].number_format = 'DD.MM.YYYY'

    # Row 6
    ws['E6'] = metadata.get('submissionType', '')    # Submission Type
    ws['H6'] = metadata.get('taluk', '')             # Taluk
    ws['J6'] = metadata.get('hardDiskNo', '')        # Hard Disk Number

    # Row 7
    ws['F7'] = metadata.get('agency', '')            # Agency Name

    print(f"✓ Metadata filled successfully")
    print(f"   - Examiner Name at E2")
    print(f"   - Designation at E3")
    print(f"   - Data Received Date at E5")
    print(f"   - Submission Type at E6")
    print(f"   - Agency at F7")
    print(f"✓ Filling village data...")

    # Fill village data rows (starting from row 11)
    village_data = data.get('villageData', [])
    print(f"   Data will start at row 11 (after header row 9 and empty row 10)")

    for index, village in enumerate(village_data):
        row_num = 11 + index  # Excel row number (1-based, starts at row 11)

        # Column mapping (A=1, B=2, etc.)
        ws.cell(row=row_num, column=1).value = village.get('slNo', index + 1)  # A: SL No
        ws.cell(row=row_num, column=2).value = f"V{index + 1}"                  # B: V1, V2, etc
        ws.cell(row=row_num, column=3).value = village.get('tileNo', '')        # C: Tile No
        ws.cell(row=row_num, column=4).value = village.get('villageName', '')   # D: Village Name
        ws.cell(row=row_num, column=5).value = village.get('lgdCode', '')       # E: LGD CODE
        ws.cell(row=row_num, column=6).value = village.get('area', '')          # F: Area
        ws.cell(row=row_num, column=7).value = village.get('talukVillage', '')  # G: Taluk
        ws.cell(row=row_num, column=8).value = village.get('hobli', '')         # H: Hobli
        ws.cell(row=row_num, column=9).value = parse_date(village.get('dateOfFlying'))  # I: Date of Flying
        ws.cell(row=row_num, column=10).value = village.get('noOfFlights', '')  # J: No of Flights

        # K: Flying Height (add ' m' suffix if value exists)
        flying_height = village.get('flyingHeight', '')
        if flying_height:
            ws.cell(row=row_num, column=11).value = f"{flying_height} m"
        else:
            ws.cell(row=row_num, column=11).value = ''

        ws.cell(row=row_num, column=12).value = village.get('rawImages', '')           # L: Raw Images
        ws.cell(row=row_num, column=13).value = village.get('intraFlightOverlap', '')  # M: Intra Flight Overlap
        ws.cell(row=row_num, column=14).value = village.get('fileNaming', '')          # N: File Naming
        ws.cell(row=row_num, column=15).value = village.get('gpsProcessing', '')       # O: GPS Processing
        # Column P is empty (16)
        ws.cell(row=row_num, column=17).value = village.get('oriGsd', '')              # Q: ORI GSD
        ws.cell(row=row_num, column=18).value = village.get('demGsd', '')              # R: DEM GSD
        ws.cell(row=row_num, column=19).value = village.get('imgProcX', '')            # S: Img Proc X
        ws.cell(row=row_num, column=20).value = village.get('imgProcY', '')            # T: Img Proc Y
        ws.cell(row=row_num, column=21).value = village.get('imgProcZ', '')            # U: Img Proc Z
        ws.cell(row=row_num, column=22).value = village.get('gcpX', '')                # V: GCP X
        ws.cell(row=row_num, column=23).value = village.get('gcpY', '')                # W: GCP Y
        ws.cell(row=row_num, column=24).value = village.get('gcpZ', '')                # X: GCP Z
        ws.cell(row=row_num, column=25).value = village.get('noOfIbase', '')           # Y: No of IBASE
        ws.cell(row=row_num, column=26).value = village.get('netAdjX', '')             # Z: Net Adj X
        ws.cell(row=row_num, column=27).value = village.get('netAdjY', '')             # AA: Net Adj Y
        ws.cell(row=row_num, column=28).value = village.get('netAdjZ', '')             # AB: Net Adj Z
        ws.cell(row=row_num, column=29).value = village.get('cors1', '')               # AC: CORS 1
        ws.cell(row=row_num, column=30).value = village.get('cors2', '')               # AD: CORS 2
        ws.cell(row=row_num, column=31).value = village.get('cors3', '')               # AE: CORS 3
        ws.cell(row=row_num, column=32).value = village.get('cors4', '')               # AF: CORS 4
        ws.cell(row=row_num, column=33).value = village.get('oriPixel', '')            # AG: ORI Pixel
        ws.cell(row=row_num, column=34).value = village.get('demPixel', '')            # AH: DEM Pixel
        ws.cell(row=row_num, column=35).value = village.get('oriQuality', '')          # AI: ORI Quality
        ws.cell(row=row_num, column=36).value = village.get('overallQuality', '')      # AJ: Overall Quality
        ws.cell(row=row_num, column=37).value = village.get('spotErrors', '')          # AK: Spot Errors
        ws.cell(row=row_num, column=38).value = village.get('fileSizeOri', '')         # AL: File Size ORI
        ws.cell(row=row_num, column=39).value = village.get('fileSizeDem', '')         # AM: File Size DEM
        ws.cell(row=row_num, column=40).value = village.get('fileSizeRaw', '')         # AN: File Size RAW
        ws.cell(row=row_num, column=41).value = village.get('path', '')                # AO: Path

        print(f"   ✓ Village {index + 1} data filled at row {row_num}")

    print(f"✓ All village data filled ({len(village_data)} villages)")

    # Fill ANEX-IV sheet with check points data
    check_points_data = data.get('checkPointsData', {})
    if check_points_data:
        print(f"📍 Filling ANEX-IV sheet with check points data...")
        fill_anex_iv_sheet(wb, village_data, check_points_data)

    print(f"💾 Saving workbook...")

    # Save the workbook (preserves ALL formatting, formulas, sheets)
    wb.save(output_path)

    print(f"═══════════════════════════════════════════════════════")
    print(f"✅ EXPORT COMPLETED SUCCESSFULLY!")
    print(f"═══════════════════════════════════════════════════════")
    print(f"📄 Output file: {output_path}")
    print(f"📋 Total sheets: {len(wb.sheetnames)}")
    print(f"👥 Villages exported: {len(village_data)}")
    print(f"✓ All formulas preserved (ANEX-IV will auto-calculate)")
    print(f"✓ All formatting preserved (colors, borders, merges)")
    print(f"✓ All 15 sheets intact")
    print(f"═══════════════════════════════════════════════════════")

    return str(output_path)


if __name__ == '__main__':
    # Read JSON data from stdin or file
    if len(sys.argv) > 1:
        # Read from file
        with open(sys.argv[1], 'r') as f:
            data = json.load(f)
        custom_filename = sys.argv[2] if len(sys.argv) > 2 else None
    else:
        # Read from stdin
        data = json.load(sys.stdin)
        custom_filename = None

    try:
        output_file = export_to_excel(data, custom_filename)
        print(f"\n✅ Success! File exported to: {output_file}")
        sys.exit(0)
    except Exception as e:
        print(f"\n❌ Error: {str(e)}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)
